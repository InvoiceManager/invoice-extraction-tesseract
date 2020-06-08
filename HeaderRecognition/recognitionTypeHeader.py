import pytesseract
import cv2
import numpy as np
import re
import openpyxl
import json


def getType(image_path):
    ################PREPROCESARE##############
    # opening - erosion followed by dilation
    def opening(img):
        kernel = np.ones((5, 5), np.uint8)
        return cv2.morphologyEx(img, cv2.MORPH_OPEN, kernel)

    # canny edge detection
    def canny(img):
        return cv2.Canny(img, 100, 200)

    # get grayscale image
    def get_grayscale(img):
        return cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # noise removal
    def remove_noise(img):
        return cv2.medianBlur(img, 5)

    # thresholding
    def thresholding(img):
        return cv2.threshold(img, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]

    # dilation
    def dilate(img):
        kernel = np.ones((5, 5), np.uint8)
        return cv2.dilate(img, kernel, iterations=1)

    # erosion
    def erode(img):
        kernel = np.ones((5, 5), np.uint8)
        return cv2.erode(img, kernel, iterations=1)

    # skew correction
    def deskew(img):
        coords = np.column_stack(np.where(img > 0))
        angle = cv2.minAreaRect(coords)[-1]
        if angle < -45:
            angle = -(90 + angle)
        else:
            angle = -angle
            (h, w) = img.shape[:2]
            center = (w // 2, h // 2)
            M = cv2.getRotationMatrix2D(center, angle, 1.0)
            rotated = cv2.warpAffine(img, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
            return rotated

    # template matching
    def match_template(img, template):
        return cv2.matchTemplate(img, template, cv2.TM_CCOEFF_NORMED)


    image = cv2.imread(image_path)
    gray = get_grayscale(image)
    thresh = thresholding(gray)
    opening = opening(gray)
    canny = canny(gray)

    cv2.imshow("image", thresh)
    cv2.waitKey(0)

    # Adding custom options
    custom_config = r'-l ron --oem 3 --psm 3'
    text = pytesseract.image_to_string(image, config=custom_config)

    if 'E.ON' in text or 'e-on' in text or 'energie electric' in text or 'energie electrica' in text or 'Energie electric' in text:
        type = "eon"
    elif "cubus" in text or "Cubus" in text:
        type = "cubus"
    else:
        type = "others"
    return type


def getContent(image_path, output_txt_path,type):
    ################PREPROCESARE##############
    # opening - erosion followed by dilation
    def opening(img):
        kernel = np.ones((5, 5), np.uint8)
        return cv2.morphologyEx(img, cv2.MORPH_OPEN, kernel)

    # canny edge detection
    def canny(img):
        return cv2.Canny(img, 100, 200)

    # get grayscale image
    def get_grayscale(img):
        return cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # noise removal
    def remove_noise(img):
        return cv2.medianBlur(img, 5)

    # thresholding
    def thresholding(img):
        return cv2.threshold(img, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]

    # dilation
    def dilate(img):
        kernel = np.ones((5, 5), np.uint8)
        return cv2.dilate(img, kernel, iterations=1)

    # erosion
    def erode(img):
        kernel = np.ones((5, 5), np.uint8)
        return cv2.erode(img, kernel, iterations=1)

    # skew correction
    def deskew(img):
        coords = np.column_stack(np.where(img > 0))
        angle = cv2.minAreaRect(coords)[-1]
        if angle < -45:
            angle = -(90 + angle)
        else:
            angle = -angle
            (h, w) = img.shape[:2]
            center = (w // 2, h // 2)
            M = cv2.getRotationMatrix2D(center, angle, 1.0)
            rotated = cv2.warpAffine(img, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
        return rotated

    # template matching
    def match_template(img, template):
        return cv2.matchTemplate(img, template, cv2.TM_CCOEFF_NORMED)

    ######## HEADER ######
    def header_cubus(output_txt):
        final_data = []

        proper_noun = '(([A-Z][A-Za-z]{1,15}[ -.])([A-Z][A-Za-z][1-5]{1,15}[ -.]*)+)[\n]'
        address = '((Bulevardul|Bd.|Str.|Str|Strada)[: .,0-9a-zA-Z]+)'
        cont = '((RO)[0-9A-Z]{22})'
        banca = '^(Banca)[ 0-9A-Za-z]+'
        nr_factura = '((nr.|nr|Nr|Nr.|Numarul|Numărul|numarul|numărul)[ :][0-9]+)'
        data = '(([0-9]{2})[/.-]([0-9]{2})[/.-][0-9]{4})'
        seria = '((Seria|seria|Serie|serie)[: -]+[A-Z0-9a-z])'

        date_factura = ''
        numar_factura = ''
        seria_factura = ''
        data_emiterii = ''
        data_scadenta = ''
        nume_furnizor = 'S.C. Cubus Arts S.R.L.'
        address_furnizor = 'Strada Morii 198 Lugojoara, jud. Timis '
        cont_furnizor = ''
        banca_furnizor = ''
        nume_cumparator = ''
        address_cumparator = ''
        cont_cumparator = ''
        banca_cumparator = ''
        denumire_produs = ''
        cantitate_produs = ''
        valoare_produs = ''
        total_factura = ''

        file1 = open(output_txt, "r")
        lines = file1.readlines()
        i = 0
        line_total = 0
        print("lines",lines)
        for line in lines:
            print(i, line)
            i = i + 1
            if i == 3 and len(line.strip()) != 0:
                numar_factura = line.split(' ')[-1].strip()
            if i == 4 and re.search(data,line):
                data_emiterii = line.split(' ')[-1].strip()
            if i == 5 and re.search(data,line):
                data_scadenta = line.split(' ')[-1].strip()
            if i == 31 and len(line.strip()) != 0:
                nume_cumparator = line.strip()
            if i == 37 and len(line.strip()) != 0:
                address_cumparator = line.split(":")[1].strip()
            if i == 33 and len(line.strip()) != 0:
                address_cumparator += address_cumparator + ' ' + line.strip()
            if i == 39 and len(line.strip()) != 0:
                banca_cumparator = line.split(':')[1].strip()
            if i == 41 and len(line.strip()) != 0:
                cont_cumparator = line.split(':')[1].strip()
            if re.search("total", line) or re.search("Total", line):
                line_total = i
        j = 0
        for line in lines:
            j = j + 1
            if j == line_total:
                if len(re.findall('([1-9][0-9]+[.,]*[0-9]*)', line)) > 0:
                    total_factura = re.findall('([1-9][0-9]+[.,]*[0-9]*)', line)[0]


        file1.close()
        final_data.append(numar_factura.strip())
        final_data.append(seria_factura.strip())
        final_data.append(data_emiterii.strip())
        final_data.append(data_scadenta.strip())
        final_data.append(nume_furnizor.strip())
        final_data.append(address_furnizor.strip())
        final_data.append(cont_furnizor.strip())
        final_data.append(banca_furnizor.strip())
        final_data.append(nume_cumparator.strip())
        final_data.append(address_cumparator.strip())
        final_data.append(cont_cumparator.strip())
        final_data.append(banca_cumparator.strip())
        final_data.append(total_factura)

        return final_data

    def header_eon(output_txt):
        final_data = []

        proper_noun = '([A-Z][A-Za-z]{1,15})[ -.]([A-Z][A-Za-z]{1,15})+'
        address = '((Bulevardul|Bd.|Str.|Str|Strada)[: .,0-9a-zA-Z]+)'
        cont = '((RO)[0-9A-Z]{22})'
        banca = '^(Banca)[ 0-9A-Za-z]+'
        nr_factura = '((nr.|nr|Nr|Nr.|Numarul|Numărul|numarul|numărul)[ :][0-9]+)'
        data = '(([0-9]{2})[/.-]([0-9]{2})[/.-][0-9]{4})'
        seria = '((Seria|seria|Serie|serie)[: -]+[A-Z0-9a-z])'

        date_factura = ''
        numar_factura = ''
        seria_factura = 'MS EON'
        data_emiterii = ''
        data_scadenta = ''
        nume_furnizor = 'E.ON Energie Romania SA'
        address_furnizor = ''
        cont_furnizor = 'RO53BRDE270SV23904012700'
        banca_furnizor = 'BRD'
        nume_cumparator = ''
        address_cumparator = ''
        cont_cumparator = ''
        banca_cumparator = ''
        denumire_produs = ''
        cantitate_produs = ''
        valoare_produs = ''
        total_factura = ''

        file1 = open(output_txt, "r")
        lines = file1.readlines()
        i = 0
        j = 0
        # print("lines", lines)

        data_invoice = 0
        name_search = 0
        line_name_cump = 0
        line_data_invoice = 0

        for line in lines:

            i = i + 1
            if name_search == 0 :
                if re.search(proper_noun,line):
                    nume_cumparator = line
                    line_name_cump = i
                    name_search = 1
            if data_invoice == 0:
                if re.search("Seria", line) or re.search("Perioada de facturare", line):
                    line_data_invoice = i

        for line in lines:
            j = j + 1
            if j == line_name_cump + 4 and len(line.strip()) != 0:
                address_cumparator = line.strip()
            if j == line_name_cump + 5 and len(line.strip()) != 0:
                address_cumparator = address_cumparator + ' ' + line.strip()
            if j == line_data_invoice + 1 and len(line.strip()) != 0:
                if len(line.split(" ")) >= 2:
                    numar_factura = line.split(" ")[2]
                if len(line.split(" ")) >= 6:
                    data_scadenta = line.split(" ")[6]
                if len(line.split(" ")) >= 5:
                    data_emiterii = line.split(" ")[5]

        file1.close()
        final_data.append(numar_factura.strip())
        final_data.append(seria_factura.strip())
        final_data.append(data_emiterii.strip())
        final_data.append(data_scadenta.strip())
        final_data.append(nume_furnizor.strip())
        final_data.append(address_furnizor.strip())
        final_data.append(cont_furnizor.strip())
        final_data.append(banca_furnizor.strip())
        final_data.append(nume_cumparator.strip())
        final_data.append(address_cumparator.strip())
        final_data.append(cont_cumparator.strip())
        final_data.append(banca_cumparator.strip())
        final_data.append(total_factura)

        return final_data

    def header_general(output_txt):
        final_data = []

        antet = ''
        file1 = open(output_txt, "r")
        lines = file1.readlines()
        for line in lines:
            antet += line

        proper_noun = '(([A-Z][A-Za-z]{1,15}[ -.])([A-Z1-9][A-Za-z]{1,15}[ -.]*)+)[\n]'
        address = '((Bulevardul|Bd.|Str.|Str|Strada)[: .,0-9a-zA-Z]+)'
        cont = '((RO)[0-9A-Z]{22})'
        banca = '^(Banca)[ 0-9A-Za-z]+'
        nr_factura = '((nr.|nr|Nr|Nr.|Numarul|Numărul|numarul|numărul)[ :][0-9]+)'
        data = '(([0-9]{2})[/.-]([0-9]{2})[/.-][0-9]{4})'
        seria = '((Seria|seria|Serie|serie)[: -]+[A-Z0-9a-z])'

        date_factura = ''
        numar_factura = ''
        seria_factura = ''
        data_emiterii = ''
        data_scadenta = ''
        nume_furnizor = ''
        date_furnizor = ''
        address_furnizor = ''
        cont_furnizor = ''
        banca_furnizor = ''
        nume_cumparator = ''
        date_cumparator = ''
        address_cumparator = ''
        cont_cumparator = ''
        banca_cumparator = ''
        denumire_produs = ''
        cantitate_produs = ''
        valoare_produs = ''
        total_factura = ''

        # luam datele facturii
        count_factura = 0
        i = 0
        line_total = 0

        for line in lines:
            i = i + 1
            print(line)
            count_factura += 1
            if re.search('factura|Factura|FACTURA|FACTURĂ|Factură', line):
                break
            if re.search("total", line) or re.search("Total", line):
                line_total = i

        j = 0
        for line in lines:
            j = j + 1
            if j == line_total:
                if len(re.findall('([1-9][0-9]+[.,]*[0-9]*)', line)) > 0:
                    total_factura = re.findall('([1-9][0-9]+[.,]*[0-9]*)', line)[0]


        with open(output_txt) as myantet:
            f_lines = myantet.readlines()[(count_factura - 2):(count_factura + 2)]
            for line in f_lines:
                date_factura += line
            if re.search(nr_factura, date_factura):
                numar_factura = re.findall(nr_factura, date_factura)[0][0]
                numar_factura = re.findall(r'([0-9]+)', numar_factura)[0]
            if re.search(seria_factura, date_factura):
                if (len(re.findall(seria, date_factura)) > 0):
                    seria_factura = re.findall(seria, date_factura)[0][0]
                    seria_factura = seria_factura.split(' ')[1]

        date = re.findall(data, antet)
        if(len(date) >= 1):
            data_emiterii = date[0][0]
        if(len(date)>= 2):
            data_scadenta = date[1][0]


        list_proper_noun = re.findall(proper_noun, antet)
        if len(list_proper_noun) >0:
            nume_furnizor = list_proper_noun[0][0]

        # unde incep datele furnizorului
        line_furnizor = 0
        for line in lines:
            line_furnizor = line_furnizor + 1
            if re.search(nume_furnizor, line):
                break

            # unde incep datele cumparatorului
            line_cumparator = 0
            i = 0

            line_address_furn = 0
            # luam datele furnizorului
            with open("final/antet.txt") as myantet:
                f_lines = myantet.readlines()[line_furnizor:(line_furnizor + 10)]
                for line in f_lines:
                    i = i + 1
                    line_cumparator = line_cumparator + 1
                    date_furnizor += line
                    if re.search(address, line):
                        address_furnizor = re.findall(address, line)[0][0]
                        line_address_furn = i
                    if re.search(cont, line):
                        cont_furnizor = re.findall(cont, line)[0][0]
                    if re.search(banca, line):
                        banca_furnizor = line
                        break
                line_cumparator = line_cumparator + line_furnizor

            # luam datele cumparatorului
            j = 0
            with open("final/antet.txt") as myantet:
                f_lines = myantet.readlines()[line_cumparator:-1]
                for line in f_lines:
                    j = j + 1
                    date_cumparator += line
                    if re.search(address, line) and j != line_address_furn:
                        address_cumparator = re.findall(address, line)[0][0]
                    if re.search(cont, line):
                        cont_cumparator = re.findall(cont, line)[0][0]
                    if re.search(banca, line):
                        banca_cumparator = line
                        break
                if(len(re.findall(proper_noun, date_cumparator)) > 0):
                    nume_cumparator = re.findall(proper_noun, date_cumparator)[0][0]

        file1.close()
        final_data.append(numar_factura)
        final_data.append(seria_factura)
        final_data.append(data_emiterii)
        final_data.append(data_scadenta)
        final_data.append(nume_furnizor)
        final_data.append(address_furnizor)
        final_data.append(cont_furnizor)
        final_data.append(banca_furnizor)
        final_data.append(nume_cumparator)
        final_data.append(address_cumparator)
        final_data.append(cont_cumparator)
        final_data.append(banca_cumparator)
        final_data.append(total_factura)

        return final_data

    def deleteContent(pfile):
        pfile.seek(0)
        pfile.truncate()

    ####################
    final_data = []

    image = cv2.imread(image_path)
    gray = get_grayscale(image)
    thresh = thresholding(gray)
    opening = opening(gray)
    canny = canny(gray)

    cv2.imshow("header", thresh)
    cv2.waitKey(0)

    # Adding custom options
    custom_config = r'-l ron --oem 3 --psm 3'
    text = pytesseract.image_to_string(image, config=custom_config)

    text.replace('\n\n', '\n').replace('\n\n', '\n')
    file = open(output_txt_path, "w+", errors="ignore")
    file.write(text)

    if type == "eon":
        final_data = header_eon(output_txt_path)
    elif type == "others":
        final_data = header_general(output_txt_path)
    elif type == "cubus":
        final_data = header_cubus(output_txt_path)

    deleteContent(file)
    # file.close()
    return final_data

def writeExcelTable(json_data, excel_path, sheet_name):
    excelfile = openpyxl.load_workbook(excel_path)
    excelfile.get_sheet_names()
    sheet1 = excelfile.get_sheet_by_name(sheet_name)
    f = open(json_data, "r")
    data = json.loads(f.read())
    for key in data:
        new_row = sheet1.max_row + 1
        list_data = data[key]
        col = 1
        for i in list_data:
            print(i)
            cell = sheet1.cell(row=new_row, column=col)
            cell.value = i
            col = col + 1
    excelfile.save(excel_path)


def writeExcelHeader(final_data, excel_path, sheet_name):
    excelfile = openpyxl.load_workbook(excel_path)
    excelfile.get_sheet_names()
    sheet1 = excelfile.get_sheet_by_name("Header-Invoices")
    new_row = sheet1.max_row + 1

    id_invoice = final_data[4].replace(" ", "_") + "-" + final_data[0].replace(" ", "_")
    cell1 = sheet1.cell(row=new_row, column=1)
    cell1.value = sheet_name

    cell1 = sheet1.cell(row=new_row, column=2)
    cell1.value = final_data[0]

    cell2 = sheet1.cell(row=new_row, column=3)
    cell2.value = final_data[1]

    cell3 = sheet1.cell(row=new_row, column=4)
    cell3.value = final_data[2]

    cell4 = sheet1.cell(row=new_row, column=5)
    cell4.value = final_data[3]

    cell5 = sheet1.cell(row=new_row, column=6)
    cell5.value = final_data[4]

    cell6 = sheet1.cell(row=new_row, column=7)
    cell6.value = final_data[5]

    cell7 = sheet1.cell(row=new_row, column=8)
    cell7.value = final_data[6]

    cell8 = sheet1.cell(row=new_row, column=9)
    cell8.value = final_data[7]

    cell9 = sheet1.cell(row=new_row, column=10)
    cell9.value = final_data[8]

    cell10 = sheet1.cell(row=new_row, column=11)
    cell10.value = final_data[9]

    cell11 = sheet1.cell(row=new_row, column=12)
    cell11.value = final_data[10]

    cell12 = sheet1.cell(row=new_row, column=13)
    cell12.value = final_data[11]

    cell12 = sheet1.cell(row=new_row, column=14)
    cell12.value = final_data[12]

    excelfile.create_sheet(id_invoice)
    excelfile.save(excel_path)

